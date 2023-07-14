import os

import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter

from src.Annotator import Annotator
from src.config import BillLabels


class AccountBook:
    def __init__(self, cfg):
        self.cfg = cfg
        self.save_path = cfg.BASECSV
        self.column_name = list(cfg.DTYPE.keys())
        database = self._read_database(cfg.BASECSV)
        self.database = database
        self.data_to_append = pd.DataFrame()
        self.annotator = Annotator()
        self.balance = self._read_balance()

    def __str__(self):
        print("==============结果：")
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)
        pd.set_option('display.max_columns', 1000)
        pd.set_option('display.max_rows', None)
        pd.set_option('max_colwidth', 30)
        pd.set_option('display.width', 1000)
        print(self.database)
        return "The AccountBook has {} rows and {} columns.".format(self.database.shape[0], self.database.shape[1])

    def dfprint(self,df):
        print('debug  print:')
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)
        pd.set_option('display.max_columns', 1000)
        pd.set_option('display.max_rows', None)
        pd.set_option('max_colwidth', 30)
        pd.set_option('display.width', 1000)
        print(df)
        print("-----------The AccountBook has {} rows and {} columns.".format(df.shape[0], df.shape[1]))

    def _read_database(self, path: str) -> pd.DataFrame:
        cols = list(range(0, len(self.cfg.DTYPE.keys())))
        df = pd.read_excel(path, sheet_name='Records', usecols=cols)
        df.columns = self.column_name
        df = df.astype(self.cfg.DTYPE)
        return df

    def _read_balance(self):
        return pd.read_excel(self.cfg.BASECSV, sheet_name='余额')

    def save_database(self, path: str):
        with pd.ExcelWriter(path) as writer:
            self.database = self.database.sort_values(by='日期时间', ascending=True)
            self.database.to_excel(writer, sheet_name='Records', index=False)
            widths = np.array(self.cfg.EXCEL_WIDTH)  # 设置excel单元格列宽
            worksheet = writer.sheets['Records']
            for i, width in enumerate(widths, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = width

            self.balance.to_excel(writer, sheet_name='余额', index=False)
            writer.sheets['余额'].column_dimensions[get_column_letter(1)].width = 22

            label_name = []
            label_value = []
            for label in BillLabels:
                label_value.append(label.value)
                label_name.append(label.name)
            df_labels = pd.DataFrame([label_name, label_value]).transpose()
            df_labels.columns = ['label.name', 'label.value']
            df_labels.to_excel(writer, sheet_name='Labels', index=False)
            writer.sheets['Labels'].column_dimensions[get_column_letter(1)].width = 50
            writer.sheets['Labels'].column_dimensions[get_column_letter(2)].width = 50
        print('最终账单导出到文件: {}'.format(path))

    def append(self):
        df  = self.data_to_append
        df.loc[df['收支'].str.contains('收入'), ['修订金额']] = df['金额']
        df.loc[df['收支'].str.contains('支出'), ['修订金额']] = df['金额'] * -1
        self.database = pd.concat([self.database, df], axis=0, ignore_index=True)


    def import_records(self, csv: str,source:str):
        print(f'[ {source} ]表格开始导入：{csv}')
        if source == '微信':
            CFG = self.cfg.WEI
        elif source == '支付宝':
            CFG = self.cfg.ZFB
        else:
            raise ValueError('暂时不支持的账单来源')

        df = pd.read_csv(csv, skiprows=range(CFG['INV_ROW']), skipfooter=0, encoding=CFG['ENCODING'], engine='python') 
        if df.columns[0] != "交易时间" :      # 检查第一列的列名
            raise ValueError('CSV文件导入错误')
        df = df.iloc[:, CFG['COLS']]  #从原表格中抓取列
        df.insert(1, '来源', source)
        df.insert(8, '修订金额', 0.0)
        df.insert(10, '小类', '')
        df.columns = self.column_name
        df['金额'] = df['金额'].map(lambda x: x.strip('¥')if isinstance(x, str) else x)
        df = df.astype(self.cfg.DTYPE)
        return df

    def data_filtering(self):
        df = self.data_to_append
        before = df.shape[0]
        for column_name, keywords in self.cfg.KEYWORD_FILTER.items():
                if column_name == '金额':
                    df = df[~df[column_name].isin(keywords)]
                else:
                    for word in keywords:
                        df = df[~df[column_name].str.contains(str(word))]  
        self.data_to_append = df
        after = df.shape[0]
        print("-----------已过滤无效数据[ {} ]行,剩余有效数据[ {} ]行.\n".format(before-after, after))

    def adjust_balance(self):
        df = self.data_to_append
        # 删除重复行
        before = df.shape[0]
        df = df.drop_duplicates()
        after = df.shape[0]
        print("-----------已删除重复数据[ {} ]行,剩余有效数据[ {} ]行.\n".format(before-after, after))
        # 计算退款金额
        df.loc[df['支付状态'].str.contains('已退款.*￥.*' ), '金额'] = df['金额'] - df['支付状态'].str.extract(r'(\d+\.\d+)', expand=False).astype(float)

        before  = len(df[df['收支'].isin(['支出', '收入'])])
        # 纳入余额宝收益
        df.loc[(df['收支'].str.contains('不计收支')) & (
                    (df['商品'].str.contains('账户结息')) |
                    (df['商品'].str.contains('余额宝.*收益发放')) |
                    (df['支付状态'].str.contains('赔付成功'))         ), ['收支']] = '收入'  

        # 支付宝走消费的银行卡尾号0139 收入
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('不计收支')) & (df['商品'].str.contains('转入')) , ['收支']] = '收入'
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('不计收支')) & (df['商品'].str.contains('转出')) , ['收支']] = '支出'

        #  微信零钱 走消费的银行卡尾号0139 支出
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('充值')) , ['收支']] = '收入'
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('提现')) , ['收支']] = '支出'

        #  微信余额宝 走消费的银行卡尾号0139 支出
        df.loc[(df['大类'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('转入')) , ['收支']] = '收入'
        df.loc[(df['大类'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('转出')) , ['收支']] = '支出'

        self.data_to_append = df
        after = len(df[df['收支'].isin(['支出', '收入'])])
        print("-----------已更新收支符号[ {} ]行.\n".format(after-before))


    def _get_record_filenames(self):
        path = self.cfg.RECORDS_DIR
        file_list = os.listdir(path)
        file_list_alipay = []
        file_list_wechat = []

        for file in file_list:
            if self.cfg.ZFB['FILE_HEADER'] in file:
                file_list_alipay.append(os.path.join(path, file))
            elif self.cfg.WEI['FILE_HEADER'] in file:
                file_list_wechat.append(os.path.join(path, file))
            elif file == '.old_records':
                continue
            else:
                raise ValueError('账单文件名不正确')

        if (len(file_list_alipay) == 0) and (len(file_list_wechat) == 0):
            raise ValueError('请检查账单文件夹RECORDS_DIR，没有账单要添加')
        return file_list_alipay, file_list_wechat


    def create_records(self):
        alipay_files, wechat_files = self._get_record_filenames()

        for alipay_csv in alipay_files:
            before = self.data_to_append.shape[0]
            df_alipay = self.import_records(alipay_csv,'支付宝')
            self.data_to_append = pd.concat([self.data_to_append, df_alipay], axis=0, ignore_index=True)
            after = self.data_to_append.shape[0]
            print("-----------已导出表格数据[ {} ]行,合计有效数据[ {} ]行.\n".format(after-before, after))

        for wechat_csv in wechat_files:
            before = self.data_to_append.shape[0]
            df_wechat = self.import_records(wechat_csv,'微信')
            self.data_to_append = pd.concat([self.data_to_append, df_wechat], axis=0, ignore_index=True)
            after = self.data_to_append.shape[0]
            print("-----------已导出表格数据[ {} ]行,合计有效数据[ {} ]行.\n".format(after-before, after))
