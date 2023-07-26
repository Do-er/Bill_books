import os
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

class BatchBills:
    def __init__(self, cfg):
        self.cfg = cfg
        self.column_name = list(cfg.DTYPE.keys())
        self.bills = pd.DataFrame()

    def dfprint(self,df,string):
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)
        pd.set_option('display.max_columns', 1000)
        pd.set_option('display.max_rows', None)
        pd.set_option('max_colwidth', 30)
        pd.set_option('display.width', 1000)
        print(f'\n ======== {string}:')
        print(df)
        print(f"-----------The DataFrame has {df.shape[0]} rows and {df.shape[1]} columns.\n")
        
    def _read_balance(self):
        return pd.read_excel(self.cfg.BASECSV, sheet_name='余额')


    def _get_record_filenames(self):
        path = self.cfg.RECORDS_DIR
        file_list = os.listdir(path)
        file_list_alipay = []
        file_list_wechat = []

        for file in file_list:
            if self.cfg.ZFB['FILE_HEADER'] in file:
                file_list_alipay.append(os.path.join(path, file))
            elif self.cfg.WEI['FILE_HEADER'] in file:
                file_list_wechat.append(os.path.join(path, file))
            elif file == '.old_records':
                continue
            else:
                raise ValueError('账单文件名不正确')
        if (len(file_list_alipay) == 0) and (len(file_list_wechat) == 0):
            raise ValueError('请检查账单文件夹RECORDS_DIR，没有账单要添加')
        return file_list_alipay, file_list_wechat


    def export_excel(self):
        cfg = self.cfg
        path = cfg.SAVECSV
        if not os.path.exists(cfg.EXPORT):
            os.makedirs(cfg.EXPORT)
        with pd.ExcelWriter(path) as excel:
            self.bills = self.bills.sort_values(by='日期时间', ascending=True)
            self.bills.to_excel(excel, sheet_name='Records', index=False)
            widths = np.array(cfg.EXCEL_WIDTH)  # 设置excel单元格列宽
            worksheet = excel.sheets['Records']
            for i, width in enumerate(widths, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = width

            # Convert the dictionary to a DataFrame
            df_labels = pd.DataFrame.from_dict(cfg.Labels)

            # Export the DataFrame to an Excel file
            df_labels.to_excel(excel, sheet_name='Labels', index=False)
            excel.sheets['Labels'].column_dimensions[get_column_letter(1)].width = 8
            # excel.sheets['Labels'].column_dimensions[get_column_letter(2)].width = 50
        print(f'最终账单导出到文件: {path}')


    def import_records(self, csv: str,source:str):
        print(f'[ {source} ]表格开始导入：{csv}')
        if source == '微信':
            CFG = self.cfg.WEI
        elif source == '支付宝':
            CFG = self.cfg.ZFB
        else:
            raise ValueError('暂时不支持的账单来源')

        df = pd.read_csv(csv, skiprows=range(CFG['INV_ROW']), skipfooter=0, encoding=CFG['ENCODING'], engine='python') 
        if df.columns[0] != "交易时间" :      # 检查第一列的列名
            raise ValueError('CSV文件导入错误')
        df = df.iloc[:, CFG['COLS']]  #从原表格中抓取列
        df.insert(1, '来源', source)
        df.insert(8, '修订金额', 0.0)
        df.insert(10, '小类', '')
        df.columns = self.column_name
        df['金额'] = df['金额'].map(lambda x: x.strip('¥')if isinstance(x, str) else x)
        df = df.astype(self.cfg.DTYPE)
        return df


    def data_filtering(self):
        df = self.bills 
        before = df.shape[0]
        for column_name, keywords in self.cfg.KEYWORD_FILTER.items():
                if column_name == '金额':
                    df = df[~df[column_name].isin(keywords)]
                else:
                    for word in keywords:
                        df = df[~df[column_name].str.contains(str(word))]  
        self.bills  = df
        after = df.shape[0]
        print(f"-----------已过滤无效数据[ {before-after} ]行,剩余有效数据[ {after} ]行.\n")


    def adjust_balance(self):
        df = self.bills 
        # 删除重复行
        before = df.shape[0]
        df = df.drop_duplicates()
        after = df.shape[0]
        print(f"-----------已删除重复数据[ {before-after} ]行,剩余有效数据[ {after} ]行.\n")

        # 计算退款金额
        df.loc[df['支付状态'].str.contains('已退款.*￥.*' ), '金额'] = df['金额'] - df['支付状态'].str.extract(r'(\d+\.\d+)', expand=False).astype(float)

        ################ 收入支出转换
        before  = len(df[df['收支'].isin(['支出', '收入'])])
        # 纳入余额宝收益
        df.loc[(df['收支'].str.contains('不计收支')) & (
                    (df['商品'].str.contains('账户结息')) |
                    (df['商品'].str.contains('余额宝.*收益发放')) |
                    (df['支付状态'].str.contains('赔付成功'))         ), ['收支']] = '收入'  

        # 支付宝走消费的银行卡尾号0139 收入
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('不计收支')) & (df['商品'].str.contains('转入')) , ['收支']] = '收入'
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('不计收支')) & (df['商品'].str.contains('转出')) , ['收支']] = '支出'

        #  微信零钱 走消费的银行卡尾号0139 支出
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('充值')) , ['收支']] = '收入'
        df.loc[(df['支付方式'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('提现')) , ['收支']] = '支出'

        #  微信余额宝 走消费的银行卡尾号0139 支出
        df.loc[(df['大类'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('转入')) , ['收支']] = '收入'
        df.loc[(df['大类'].str.contains('0139')) & (df['收支'].str.contains('/')) & (df['大类'].str.contains('转出')) , ['收支']] = '支出'
        ################ 收入支出转换
        after = len(df[df['收支'].isin(['支出', '收入'])])

        ################ 转换符号到 修订金额 列
        df.loc[df['收支'].str.contains('收入'), ['修订金额']] = df['金额']
        df.loc[df['收支'].str.contains('支出'), ['修订金额']] = df['金额'] * -1
        print(f"-----------已更新收支符号[ {after-before} ]行.\n")
        self.bills  = df


    def create_records(self):
        df =  self.bills
        alipay_files, wechat_files = self._get_record_filenames()

        for alipay_csv in alipay_files:
            before = df.shape[0]
            df_alipay = self.import_records(alipay_csv,'支付宝')
            df = pd.concat([df, df_alipay], axis=0, ignore_index=True)
            after = df.shape[0]
            print(f"-----------已导出表格数据[ {after-before} ]行,合计有效数据[ {after} ]行.\n")

        for wechat_csv in wechat_files:
            before = df.shape[0]
            df_wechat = self.import_records(wechat_csv,'微信')
            df = pd.concat([df, df_wechat], axis=0, ignore_index=True)
            after = df.shape[0]
            print(f"-----------已导出表格数据[ {after-before} ]行,合计有效数据[ {after} ]行.\n")
        self.bills = df
